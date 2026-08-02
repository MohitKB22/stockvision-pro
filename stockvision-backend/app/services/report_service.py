"""
Report Generation Service — new in v2.0.

Produces real, downloadable artifacts (PDF via ReportLab, CSV via the stdlib,
XLSX via openpyxl) for the four report types in the reference design. The Reports
page previously had buttons wired to nothing.

Design decision: every report is built from the SAME service calls the UI renders
from (PortfolioService, RiskAnalyticsService, the ML registry). A report that
recomputes its numbers through a separate path is a report that eventually
disagrees with the screen it was generated from — the most common and most
damaging class of bug in financial reporting tools.
"""
import csv
import io
import logging
import uuid
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.exceptions import BadRequestException, NotFoundException
from app.domain.enums import AuditAction, ReportFormat, ReportType
from app.domain.markets import get_market
from app.models.system import GeneratedReport
from app.repositories.portfolio_repository import PortfolioRepository
from app.repositories.system_repository import ReportRepository
from app.services.audit_service import AuditService
from app.services.portfolio_service import PortfolioService
from app.services.risk_analytics_service import RiskAnalyticsService

logger = logging.getLogger(__name__)

# Brand palette, matched to the frontend design tokens so an exported PDF is
# recognisably the same product as the screen it came from.
_INK = colors.HexColor("#0B0F1A")
_ACCENT = colors.HexColor("#3B82F6")
_MUTED = colors.HexColor("#64748B")
_ROW_ALT = colors.HexColor("#F1F5F9")


class ReportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.portfolios = PortfolioRepository(db)
        self.portfolio_service = PortfolioService(db)
        self.risk = RiskAnalyticsService(db)
        self.repo = ReportRepository(db)
        self.audit = AuditService(db)

    # --- Entry point ----------------------------------------------------------
    def generate(
        self,
        report_type: ReportType,
        report_format: ReportFormat,
        portfolio_id: uuid.UUID | None = None,
        lookback_days: int = 252,
    ) -> GeneratedReport:
        if report_type in {ReportType.PORTFOLIO, ReportType.RISK, ReportType.TAX} and portfolio_id is None:
            portfolio = self.portfolios.get_default()
            if portfolio is None:
                raise BadRequestException(
                    "A portfolio is required for this report type, and none exist yet."
                )
            portfolio_id = portfolio.id

        sections = self._build_sections(report_type, portfolio_id, lookback_days)

        storage_dir = Path(settings.REPORT_STORAGE_DIR)
        storage_dir.mkdir(parents=True, exist_ok=True)
        report_id = uuid.uuid4()
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        filename = f"{report_type.value}-report-{stamp}.{report_format.extension}"
        path = storage_dir / f"{report_id}_{filename}"

        renderers = {
            ReportFormat.PDF: self._render_pdf,
            ReportFormat.CSV: self._render_csv,
            ReportFormat.EXCEL: self._render_xlsx,
        }
        payload = renderers[report_format](sections)
        path.write_bytes(payload)

        record = self.repo.create(
            GeneratedReport(
                id=report_id, report_type=report_type, report_format=report_format,
                title=sections["title"], portfolio_id=portfolio_id,
                storage_path=str(path), size_bytes=len(payload),
                parameters={"lookback_days": lookback_days, "filename": filename},
            )
        )
        self.audit.log(
            action=AuditAction.REPORT_GENERATED,
            resource=f"report:{record.id}",
            detail={"type": report_type.value, "format": report_format.value},
        )
        return record

    def read_file(self, report_id: uuid.UUID) -> tuple[bytes, str, str]:
        record = self.repo.get(report_id)
        if not record:
            raise NotFoundException("Report not found")
        path = Path(record.storage_path)
        if not path.exists():
            raise NotFoundException("The generated file for this report is no longer on disk.")
        filename = (record.parameters or {}).get("filename") or path.name
        return path.read_bytes(), filename, ReportFormat(record.report_format).media_type

    # --- Section builders -------------------------------------------------------
    def _build_sections(
        self, report_type: ReportType, portfolio_id: uuid.UUID | None, lookback_days: int
    ) -> dict:
        builders = {
            ReportType.PORTFOLIO: self._portfolio_sections,
            ReportType.RISK: self._risk_sections,
            ReportType.PREDICTION: self._prediction_sections,
            ReportType.TAX: self._tax_sections,
        }
        return builders[report_type](portfolio_id, lookback_days)

    def _portfolio_sections(self, portfolio_id: uuid.UUID, lookback_days: int) -> dict:
        summary = self.portfolio_service.get_summary(portfolio_id)
        sym = get_market(summary.market).currency_symbol
        return {
            "title": f"Portfolio Report — {summary.name}",
            "subtitle": f"{summary.holding_count} holdings · benchmark {summary.benchmark_symbol}",
            "metrics": [
                ("Total Value", f"{sym}{summary.total_value:,.2f}"),
                ("Invested Cost", f"{sym}{summary.total_cost_basis:,.2f}"),
                ("Unrealized P&L", f"{sym}{summary.total_unrealized_pnl:,.2f} ({summary.total_unrealized_pnl_pct:+.2%})"),
                ("Realized P&L", f"{sym}{summary.total_realized_pnl:,.2f}"),
                ("Day Change", f"{sym}{summary.day_change:,.2f} ({summary.day_change_pct:+.2%})"),
                ("Cash Balance", f"{sym}{summary.cash_balance:,.2f}"),
            ],
            "tables": [
                {
                    "name": "Holdings",
                    "columns": ["Symbol", "Name", "Sector", "Qty", "Avg Cost", "Price", "Value", "P&L", "P&L %", "Weight"],
                    "rows": [
                        [h.symbol, h.name, h.sector or "—", f"{h.quantity:g}",
                         f"{h.average_cost:,.2f}", f"{h.current_price:,.2f}",
                         f"{h.market_value:,.2f}", f"{h.unrealized_pnl:,.2f}",
                         f"{h.unrealized_pnl_pct:+.2%}", f"{h.weight_pct:.2%}"]
                        for h in summary.holdings
                    ],
                },
                {
                    "name": "Sector Allocation",
                    "columns": ["Sector", "Value", "Weight"],
                    "rows": [[s.label, f"{s.value:,.2f}", f"{s.weight_pct:.2%}"] for s in summary.sector_exposure],
                },
            ],
        }

    def _risk_sections(self, portfolio_id: uuid.UUID, lookback_days: int) -> dict:
        metrics = self.risk.metrics(portfolio_id, lookback_days)
        stress = self.risk.stress(portfolio_id, lookback_days)
        correlations = self.risk.correlations(portfolio_id, lookback_days)

        tables = [
            {
                "name": "Stress Scenarios",
                "columns": ["Scenario", "Market Shock", "Portfolio Impact", "Impact Value", "Resulting Value"],
                # Attribute access, not subscripting: RiskAnalyticsService returns
                # a validated Pydantic StressTestResponse, not raw dicts. Indexing
                # these was a real TypeError at runtime — caught by the end-to-end
                # smoke test, which is why that test generates all four report
                # types in all three formats rather than spot-checking one.
                "rows": [
                    [s.scenario, f"{s.market_shock_pct:+.2%}", f"{s.portfolio_impact_pct:+.2%}",
                     f"{s.portfolio_impact_value:,.2f}", f"{s.resulting_value:,.2f}"]
                    for s in stress.scenarios
                ],
            }
        ]
        if correlations.matrix:
            tables.append({
                "name": "Correlation Matrix",
                "columns": ["", *correlations.labels],
                "rows": [
                    [correlations.labels[i], *[f"{v:.2f}" for v in row]]
                    for i, row in enumerate(correlations.matrix)
                ],
            })

        return {
            "title": "Risk Report",
            "subtitle": f"{metrics.observations} observations · {lookback_days}-day lookback · benchmark {metrics.benchmark_symbol}",
            "metrics": [
                ("Annualized Return", f"{metrics.annualized_return:+.2%}"),
                ("Annualized Volatility", f"{metrics.annualized_volatility:.2%}"),
                ("Sharpe Ratio", f"{metrics.sharpe_ratio:.2f}"),
                ("Sortino Ratio", f"{metrics.sortino_ratio:.2f}"),
                ("Max Drawdown", f"{metrics.max_drawdown:.2%}"),
                ("VaR 95% (Historical)", f"{metrics.value_at_risk_95_historical:.2%}"),
                ("VaR 95% (Parametric)", f"{metrics.value_at_risk_95_parametric:.2%}"),
                ("VaR 95% (Monte Carlo)", f"{metrics.value_at_risk_95_monte_carlo:.2%}"),
                ("Expected Shortfall 95%", f"{metrics.expected_shortfall_95:.2%}"),
                ("Beta", f"{metrics.beta:.2f}" if metrics.beta is not None else "n/a"),
                ("Alpha (annualized)", f"{metrics.alpha:+.2%}" if metrics.alpha is not None else "n/a"),
            ],
            "tables": tables,
        }

    def _prediction_sections(self, portfolio_id: uuid.UUID | None, lookback_days: int) -> dict:
        from app.repositories.ml_repository import MLModelRepository, SignalRepository

        models = MLModelRepository(self.db).list_models(limit=50)
        signals = SignalRepository(self.db).list_recent(limit=50)

        return {
            "title": "AI Prediction Report",
            "subtitle": f"{len(models)} registered model versions · {len(signals)} recent signals",
            "metrics": [
                ("Registered Models", str(len(models))),
                ("Production Models", str(sum(1 for m in models if str(m.stage) == "production"))),
                ("Recent Signals", str(len(signals))),
                (
                    "Mean Signal Confidence",
                    f"{(sum(s.confidence for s in signals) / len(signals)):.1%}" if signals else "n/a",
                ),
            ],
            "tables": [
                {
                    "name": "Model Registry",
                    "columns": ["Model", "Version", "Task", "Algorithm", "Stage", "Accuracy", "F1", "Trained"],
                    "rows": [
                        [m.name, str(m.version), str(m.task), str(m.algorithm), str(m.stage),
                         f"{(m.metrics or {}).get('accuracy', 0) or 0:.3f}",
                         f"{(m.metrics or {}).get('f1', 0) or 0:.3f}",
                         m.trained_at.strftime("%Y-%m-%d %H:%M")]
                        for m in models
                    ],
                },
                {
                    "name": "Recent Signals",
                    "columns": ["Symbol", "Action", "Confidence", "Risk", "Generated"],
                    "rows": [
                        [s.stock.symbol if s.stock else "—", str(s.action).replace("_", " ").upper(),
                         f"{s.confidence:.1%}", f"{s.risk_score:.2f}",
                         s.created_at.strftime("%Y-%m-%d %H:%M")]
                        for s in signals
                    ],
                },
            ],
        }

    def _tax_sections(self, portfolio_id: uuid.UUID, lookback_days: int) -> dict:
        """
        Capital-gains summary built from the order ledger.

        Explicitly NOT tax advice, and the document says so: holding-period
        rules, rates and offset treatment differ by jurisdiction. What this
        produces is the underlying disposal data an accountant needs, not a
        computed liability the software has no business asserting.
        """
        summary = self.portfolio_service.get_summary(portfolio_id)
        transactions = self.portfolio_service.get_transactions(portfolio_id, limit=1000)
        sym = get_market(summary.market).currency_symbol

        buys = [t for t in transactions if str(t.side) == "buy"]
        sells = [t for t in transactions if str(t.side) == "sell"]
        total_costs = sum(t.transaction_cost + t.slippage for t in transactions)

        return {
            "title": "Tax Report — Capital Gains Summary",
            "subtitle": f"{summary.name} · {len(transactions)} transactions on record",
            "disclaimer": (
                "This document summarises transaction and disposal data only. It is not tax "
                "advice and does not compute a tax liability: holding-period thresholds, "
                "applicable rates, indexation and loss-offset rules vary by jurisdiction and "
                "by taxpayer. Provide this summary to a qualified tax professional."
            ),
            "metrics": [
                ("Realized P&L", f"{sym}{summary.total_realized_pnl:,.2f}"),
                ("Unrealized P&L", f"{sym}{summary.total_unrealized_pnl:,.2f}"),
                ("Purchases", str(len(buys))),
                ("Disposals", str(len(sells))),
                ("Transaction Costs", f"{sym}{total_costs:,.2f}"),
            ],
            "tables": [
                {
                    "name": "Transactions",
                    "columns": ["Date", "Symbol", "Side", "Qty", "Price", "Value", "Costs"],
                    "rows": [
                        [t.executed_at.strftime("%Y-%m-%d"), t.symbol, str(t.side).upper(),
                         f"{t.quantity:g}", f"{t.price:,.2f}", f"{t.value:,.2f}",
                         f"{t.transaction_cost + t.slippage:,.2f}"]
                        for t in transactions
                    ],
                },
            ],
        }

    # --- Renderers -----------------------------------------------------------------
    def _render_pdf(self, sections: dict) -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=16 * mm, rightMargin=16 * mm, topMargin=18 * mm, bottomMargin=16 * mm,
            title=sections["title"], author="StockVision Pro",
        )
        base = getSampleStyleSheet()
        h1 = ParagraphStyle("H1", parent=base["Heading1"], fontSize=19, textColor=_INK, spaceAfter=2)
        sub = ParagraphStyle("Sub", parent=base["Normal"], fontSize=9.5, textColor=_MUTED, spaceAfter=12)
        h2 = ParagraphStyle("H2", parent=base["Heading2"], fontSize=12.5, textColor=_ACCENT, spaceBefore=14, spaceAfter=6)
        note = ParagraphStyle("Note", parent=base["Normal"], fontSize=8, textColor=_MUTED, leading=11)

        story = [Paragraph(sections["title"], h1), Paragraph(sections.get("subtitle", ""), sub)]

        if metrics := sections.get("metrics"):
            story.append(Paragraph("Key Metrics", h2))
            table = Table([list(pair) for pair in metrics], colWidths=[70 * mm, 100 * mm], hAlign="LEFT")
            table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TEXTCOLOR", (0, 0), (0, -1), _MUTED),
                ("TEXTCOLOR", (1, 0), (1, -1), _INK),
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, _ROW_ALT]),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ]))
            story.append(table)

        for spec in sections.get("tables") or []:
            if not spec["rows"]:
                continue
            story.append(Paragraph(spec["name"], h2))
            # Cap rows per table so a 1,000-row ledger doesn't produce a 60-page
            # PDF; the CSV/XLSX exports carry the complete data.
            rows = spec["rows"][:60]
            table = Table([spec["columns"], *rows], repeatRows=1, hAlign="LEFT")
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), _INK),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.2),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _ROW_ALT]),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(table)
            if len(spec["rows"]) > len(rows):
                story.append(Spacer(1, 4))
                story.append(Paragraph(
                    f"Showing {len(rows)} of {len(spec['rows'])} rows. "
                    "Export as CSV or Excel for the complete dataset.", note,
                ))

        if sections.get("disclaimer"):
            story.append(PageBreak())
            story.append(Paragraph("Important Notice", h2))
            story.append(Paragraph(sections["disclaimer"], note))

        story.append(Spacer(1, 14))
        story.append(Paragraph(
            f"Generated by StockVision Pro v{settings.APP_VERSION} on "
            f"{datetime.now(timezone.utc).strftime('%d %b %Y at %H:%M UTC')}. "
            "Figures are derived from the platform's stored market data and are for "
            "analysis purposes only — not investment advice.", note,
        ))

        doc.build(story)
        return buffer.getvalue()

    def _render_csv(self, sections: dict) -> bytes:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([sections["title"]])
        writer.writerow([sections.get("subtitle", "")])
        writer.writerow([])
        writer.writerow(["Key Metrics"])
        for label, value in sections.get("metrics") or []:
            writer.writerow([label, value])

        for spec in sections.get("tables") or []:
            writer.writerow([])
            writer.writerow([spec["name"]])
            writer.writerow(spec["columns"])
            writer.writerows(spec["rows"])

        if sections.get("disclaimer"):
            writer.writerow([])
            writer.writerow(["Notice", sections["disclaimer"]])

        # utf-8-sig: Excel on Windows renders a plain UTF-8 CSV containing ₹ as
        # mojibake unless a BOM is present. One byte sequence removes an entire
        # category of "your export is broken" report.
        return buffer.getvalue().encode("utf-8-sig")

    def _render_xlsx(self, sections: dict) -> bytes:
        workbook = Workbook()
        header_fill = PatternFill("solid", fgColor="0B0F1A")
        header_font = Font(color="FFFFFF", bold=True, size=10)

        overview = workbook.active
        overview.title = "Overview"
        overview["A1"] = sections["title"]
        overview["A1"].font = Font(bold=True, size=14)
        overview["A2"] = sections.get("subtitle", "")
        overview["A2"].font = Font(color="64748B", size=10)
        overview["A4"], overview["B4"] = "Metric", "Value"
        for cell in ("A4", "B4"):
            overview[cell].fill = header_fill
            overview[cell].font = header_font
        for offset, (label, value) in enumerate(sections.get("metrics") or [], start=5):
            overview[f"A{offset}"] = label
            overview[f"B{offset}"] = value
        overview.column_dimensions["A"].width = 34
        overview.column_dimensions["B"].width = 42

        for spec in sections.get("tables") or []:
            # Excel sheet names cap at 31 chars and forbid []:*?/\
            safe = "".join(c for c in spec["name"] if c not in "[]:*?/\\")[:31] or "Sheet"
            sheet = workbook.create_sheet(safe)
            sheet.append(spec["columns"])
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for row in spec["rows"]:
                sheet.append(row)
            sheet.freeze_panes = "A2"
            for index, column_name in enumerate(spec["columns"], start=1):
                widths = [len(str(column_name))]
                widths += [len(str(r[index - 1])) for r in spec["rows"][:400] if len(r) >= index]
                sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = min(
                    max(max(widths) + 2, 10), 48
                )

        if sections.get("disclaimer"):
            notice = workbook.create_sheet("Notice")
            notice["A1"] = "Important Notice"
            notice["A1"].font = Font(bold=True, size=12)
            notice["A3"] = sections["disclaimer"]
            notice["A3"].alignment = Alignment(wrap_text=True, vertical="top")
            notice.column_dimensions["A"].width = 110

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
